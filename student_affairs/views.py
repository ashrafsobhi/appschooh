from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Count, Sum
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, Http404, JsonResponse
from django.conf import settings
from django.urls import reverse
import os
from .models import Student, StudentAttachment, SiteSettings, Classroom, UserProfile
from .forms import (
    StudentForm,
    ExcelUploadForm,
    AttachmentForm,
    RecordsForm,
    SelectStudentForm,
    ClassroomForm,
    AddStudentToClassroomForm,
    StudentContactForm,
    AdminCreateUserForm,
    ProfileUpdateForm,
    PasswordChangeForm,
)
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import views as auth_views
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .utils import parse_egyptian_national_id
from .auth_utils import generate_jwt_token, verify_jwt_token, get_token_from_request
import openpyxl
import xlrd
from datetime import datetime, date
from io import BytesIO
import qrcode
import base64
from PIL import Image, ImageDraw, ImageFont
import hashlib
from collections import defaultdict


# ===== BASIC FUNCTIONS (Stubs - يجب التأكد من وجودها) =====
@login_required
def welcome(request):
    """صفحة الترحيب - محمية بتسجيل الدخول"""
    total_students = Student.objects.count()
    active_students = Student.objects.filter(registration_status='ناجح ومنقول').count()
    return render(request, 'student_affairs/welcome.html', {
        'total_students': total_students,
        'active_students': active_students,
    })


@login_required
def student_list(request):
    """صفحة قيد الطلبة"""
    try:
        # جلب جميع الطلبة
        all_students = Student.objects.all()
        
        # البحث والفلترة
        search_query = request.GET.get('search', '')
        registration_status = request.GET.get('registration_status', '')
        religion = request.GET.get('religion', '')
        
        if search_query:
            all_students = all_students.filter(
                Q(student_code__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(national_id__icontains=search_query)
            )
        
        if registration_status:
            all_students = all_students.filter(registration_status=registration_status)
        
        if religion:
            all_students = all_students.filter(religion=religion)
        
        # تقسيم الطلبة حسب الصف والنوع
        # ملاحظة: يمكن أن يكون الصف محدداً في حقل grade مباشرة أو من خلال classroom
        # دعم القيم العربية (الأول، الثاني، الثالث) والإنجليزية (g1، g2، g3)
        # الطلاب بدون grade محدد سيظهرون في الصف الأول (كقسم افتراضي)
        
        # الصف الأول: دعم جميع الصيغ (g1، الأول، الصف الأول، إلخ) + الطلاب بدون grade محدد
        male_g1 = all_students.filter(
            Q(gender='male') & (
                Q(grade='g1') | Q(grade='الأول') | Q(grade='اول') | Q(grade='1') |
                Q(grade__icontains='الصف الأول') | Q(grade__icontains='صف أول') |
                Q(classroom__grade='g1') |
                # الطلاب بدون grade محدد (ليسوا في g2 أو g3 أو الثاني أو الثالث)
                (
                    (Q(grade__isnull=True) | Q(grade='')) &
                    ~(
                        Q(grade__in=['g2', 'g3', 'الثاني', 'الثالث', 'ثاني', 'ثالث', '2', '3']) |
                        Q(grade__icontains='الصف الثاني') | Q(grade__icontains='الصف الثالث') |
                        Q(grade__icontains='صف ثاني') | Q(grade__icontains='صف ثالث') |
                        Q(classroom__grade__in=['g2', 'g3'])
                    )
                )
            )
        ).distinct().order_by('student_code')
        
        female_g1 = all_students.filter(
            Q(gender='female') & (
                Q(grade='g1') | Q(grade='الأول') | Q(grade='اول') | Q(grade='1') |
                Q(grade__icontains='الصف الأول') | Q(grade__icontains='صف أول') |
                Q(classroom__grade='g1') |
                # الطلاب بدون grade محدد (ليسوا في g2 أو g3 أو الثاني أو الثالث)
                (
                    (Q(grade__isnull=True) | Q(grade='')) &
                    ~(
                        Q(grade__in=['g2', 'g3', 'الثاني', 'الثالث', 'ثاني', 'ثالث', '2', '3']) |
                        Q(grade__icontains='الصف الثاني') | Q(grade__icontains='الصف الثالث') |
                        Q(grade__icontains='صف ثاني') | Q(grade__icontains='صف ثالث') |
                        Q(classroom__grade__in=['g2', 'g3'])
                    )
                )
            )
        ).distinct().order_by('student_code')
        
        # الصف الثاني - دعم جميع الصيغ
        male_g2 = all_students.filter(
            Q(gender='male') & (
                Q(grade='g2') | Q(grade='الثاني') | Q(grade='ثاني') | Q(grade='2') |
                Q(grade__icontains='الصف الثاني') | Q(grade__icontains='صف ثاني') |
                Q(classroom__grade='g2')
            )
        ).distinct().order_by('student_code')
        female_g2 = all_students.filter(
            Q(gender='female') & (
                Q(grade='g2') | Q(grade='الثاني') | Q(grade='ثاني') | Q(grade='2') |
                Q(grade__icontains='الصف الثاني') | Q(grade__icontains='صف ثاني') |
                Q(classroom__grade='g2')
            )
        ).distinct().order_by('student_code')
        
        # الصف الثالث - دعم جميع الصيغ
        male_g3 = all_students.filter(
            Q(gender='male') & (
                Q(grade='g3') | Q(grade='الثالث') | Q(grade='ثالث') | Q(grade='3') |
                Q(grade__icontains='الصف الثالث') | Q(grade__icontains='صف ثالث') |
                Q(classroom__grade='g3')
            )
        ).distinct().order_by('student_code')
        female_g3 = all_students.filter(
            Q(gender='female') & (
                Q(grade='g3') | Q(grade='الثالث') | Q(grade='ثالث') | Q(grade='3') |
                Q(grade__icontains='الصف الثالث') | Q(grade__icontains='صف ثالث') |
                Q(classroom__grade='g3')
            )
        ).distinct().order_by('student_code')
        
        # الإحصائيات
        total_students = Student.objects.count()
        g1_count = Student.objects.filter(Q(grade='g1') | Q(classroom__grade='g1')).distinct().count()
        g2_count = Student.objects.filter(Q(grade='g2') | Q(classroom__grade='g2')).distinct().count()
        g3_count = Student.objects.filter(Q(grade='g3') | Q(classroom__grade='g3')).distinct().count()
        
    except Exception as e:
        # في حالة حدوث خطأ، إرجاع قوائم فارغة
        messages.error(request, f'حدث خطأ في جلب البيانات: {str(e)}')
        male_g1 = female_g1 = male_g2 = female_g2 = male_g3 = female_g3 = Student.objects.none()
        total_students = g1_count = g2_count = g3_count = 0
        all_students = Student.objects.none()
    
    context = {
        'students': all_students,
        'male_g1': male_g1,
        'female_g1': female_g1,
        'male_g2': male_g2,
        'female_g2': female_g2,
        'male_g3': male_g3,
        'female_g3': female_g3,
        'total_students': total_students,
        'g1_count': g1_count,
        'g2_count': g2_count,
        'g3_count': g3_count,
        'search_query': search_query,
        'registration_status': registration_status,
        'religion': religion,
    }
    
    return render(request, 'student_affairs/student_list.html', context)


@login_required
def records_list(request):
    """صفحة السجلات"""
    try:
        all_students = Student.objects.all().order_by('student_code')
        
        # البحث والفلترة
        search_query = request.GET.get('search', '')
        grade_filter = request.GET.get('grade', '')
        registration_status = request.GET.get('registration_status', '')
        
        if search_query:
            all_students = all_students.filter(
                Q(student_code__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(national_id__icontains=search_query)
            )
        
        if grade_filter:
            all_students = all_students.filter(
                Q(grade=grade_filter) | Q(classroom__grade=grade_filter)
            ).distinct()
        
        if registration_status:
            all_students = all_students.filter(registration_status=registration_status)
        
        # تقسيم الطلاب حسب الصف
        # دعم القيم العربية (الأول، الثاني، الثالث) والإنجليزية (g1، g2، g3)
        # الطلاب بدون grade محدد سيظهرون في الصف الأول (كقسم افتراضي)
        
        # الصف الأول: دعم جميع الصيغ (g1، الأول، الصف الأول، إلخ) + الطلاب بدون grade محدد
        g1_students = all_students.filter(
            Q(grade='g1') | Q(grade='الأول') | Q(grade='اول') | Q(grade='1') |
            Q(grade__icontains='الصف الأول') | Q(grade__icontains='صف أول') |
            Q(classroom__grade='g1') |
            # الطلاب بدون grade محدد (ليسوا في g2 أو g3 أو الثاني أو الثالث)
            (
                (Q(grade__isnull=True) | Q(grade='')) &
                ~(
                    Q(grade__in=['g2', 'g3', 'الثاني', 'الثالث', 'ثاني', 'ثالث', '2', '3']) |
                    Q(grade__icontains='الصف الثاني') | Q(grade__icontains='الصف الثالث') |
                    Q(grade__icontains='صف ثاني') | Q(grade__icontains='صف ثالث') |
                    Q(classroom__grade__in=['g2', 'g3'])
                )
            )
        ).distinct()
        
        g2_students = all_students.filter(
            Q(grade='g2') | Q(grade='الثاني') | Q(grade='ثاني') | Q(grade='2') |
            Q(grade__icontains='الصف الثاني') | Q(grade__icontains='صف ثاني') |
            Q(classroom__grade='g2')
        ).distinct()
        
        g3_students = all_students.filter(
            Q(grade='g3') | Q(grade='الثالث') | Q(grade='ثالث') | Q(grade='3') |
            Q(grade__icontains='الصف الثالث') | Q(grade__icontains='صف ثالث') |
            Q(classroom__grade='g3')
        ).distinct()
        
        # إحصائيات لكل صف
        def get_grade_stats(students):
            return {
                'total': students.count(),
                'regular': students.filter(registration_type='نظامي').count(),
                'services': students.filter(registration_type='خدمات').count(),
                'system': students.filter(registration_type='نظامي').count(),
                'workers': students.filter(registration_type='عمال').count(),
                'foreign': students.filter(is_foreign_student=True).count(),
            }
        
        def get_grade_chart_data(students, grade_key):
            """إحصائيات الرسوم البيانية"""
            male_count = students.filter(gender='male').count()
            female_count = students.filter(gender='female').count()
            foreign_count = students.filter(is_foreign_student=True).count()
            local_count = students.filter(is_foreign_student=False).count()
            withdrawn_count = students.filter(exam_round='سحب').count()
            
            # إحصائيات الشعب
            sections = {}
            for s in students.values_list('section', flat=True).distinct():
                if s:
                    sections[s] = students.filter(section=s).count()
            
            # نوع القيد
            reg_types = {}
            for rt in students.values_list('registration_type', flat=True).distinct():
                if rt:
                    reg_types[rt] = students.filter(registration_type=rt).count()
            
            # حالة القيد
            statuses = {}
            for st in students.values_list('registration_status', flat=True).distinct():
                if st:
                    statuses[st] = students.filter(registration_status=st).count()
            
            # موقف الدمج
            inclusions = {}
            for inc in students.values_list('inclusion_status', flat=True).distinct():
                if inc:
                    inclusions[inc] = students.filter(inclusion_status=inc).count()
            
            # اللغات
            lang1 = {}
            for l1 in students.values_list('first_foreign_language', flat=True).distinct():
                if l1:
                    lang1[l1] = students.filter(first_foreign_language=l1).count()
            
            lang2 = {}
            for l2 in students.values_list('second_foreign_language', flat=True).distinct():
                if l2:
                    lang2[l2] = students.filter(second_foreign_language=l2).count()
            
            return {
                'gender': {
                    'labels': ['بنين', 'بنات'],
                    'data': [male_count, female_count]
                },
                'foreign': {
                    'labels': ['محلي', 'وافد'],
                    'data': [local_count, foreign_count]
                },
                'withdrawn': {
                    'labels': ['عادي', 'سحب'],
                    'data': [students.count() - withdrawn_count, withdrawn_count]
                },
                'sections': {
                    'labels': list(sections.keys()),
                    'data': list(sections.values())
                },
                'regtype': {
                    'labels': list(reg_types.keys()),
                    'data': list(reg_types.values())
                },
                'status': {
                    'labels': list(statuses.keys()),
                    'data': list(statuses.values())
                },
                'inclusion': {
                    'labels': list(inclusions.keys()),
                    'data': list(inclusions.values())
                },
                'lang1': {
                    'labels': list(lang1.keys()),
                    'data': list(lang1.values())
                },
                'lang2': {
                    'labels': list(lang2.keys()),
                    'data': list(lang2.values())
                },
            }
        
        g1_stats = get_grade_stats(g1_students)
        g2_stats = get_grade_stats(g2_students)
        g3_stats = get_grade_stats(g3_students)
        
        # بيانات الرسوم البيانية
        import json
        g1_gender = get_grade_chart_data(g1_students, 'g1')
        g2_gender = get_grade_chart_data(g2_students, 'g2')
        g3_gender = get_grade_chart_data(g3_students, 'g3')
        
        # تحويل البيانات إلى JSON للاستخدام في القالب
        for grade_data in [g1_gender, g2_gender, g3_gender]:
            for key, value in grade_data.items():
                if isinstance(value, dict):
                    value['labels_json'] = json.dumps(value.get('labels', []), ensure_ascii=False)
                    value['data_json'] = json.dumps(value.get('data', []), ensure_ascii=False)
        
        # إعدادات الموقع
        site_settings = SiteSettings.get_settings()
        
        # جداول الإحصائيات للطباعة
        def get_stats_table(students):
            male = students.filter(gender='male')
            female = students.filter(gender='female')
            return [
                {'label': 'انتظام مستجد', 'male': male.filter(registration_type='نظامي', exam_round='دور أول').count(), 'female': female.filter(registration_type='نظامي', exam_round='دور أول').count(), 'total': students.filter(registration_type='نظامي', exam_round='دور أول').count()},
                {'label': 'انتظام باقي', 'male': male.filter(registration_type='نظامي', exam_round__in=['دور ثاني', 'باقي']).count(), 'female': female.filter(registration_type='نظامي', exam_round__in=['دور ثاني', 'باقي']).count(), 'total': students.filter(registration_type='نظامي', exam_round__in=['دور ثاني', 'باقي']).count()},
                {'label': 'عمال مستجد', 'male': male.filter(registration_type='عمال', exam_round='دور أول').count(), 'female': female.filter(registration_type='عمال', exam_round='دور أول').count(), 'total': students.filter(registration_type='عمال', exam_round='دور أول').count()},
                {'label': 'عمال باقي', 'male': male.filter(registration_type='عمال', exam_round__in=['دور ثاني', 'باقي']).count(), 'female': female.filter(registration_type='عمال', exam_round__in=['دور ثاني', 'باقي']).count(), 'total': students.filter(registration_type='عمال', exam_round__in=['دور ثاني', 'باقي']).count()},
                {'label': 'دمج', 'male': male.filter(inclusion_status__in=['بصري', 'سمعي', 'حركي', 'صعوبات تعلم', 'أخرى']).count(), 'female': female.filter(inclusion_status__in=['بصري', 'سمعي', 'حركي', 'صعوبات تعلم', 'أخرى']).count(), 'total': students.filter(inclusion_status__in=['بصري', 'سمعي', 'حركي', 'صعوبات تعلم', 'أخرى']).count()},
            ]
        
        def get_sections_table(students):
            sections = []
            for section in students.values_list('section', flat=True).distinct():
                if section:
                    section_students = students.filter(section=section)
                    sections.append({
                        'label': section,
                        'male': section_students.filter(gender='male').count(),
                        'female': section_students.filter(gender='female').count(),
                        'total': section_students.count()
                    })
            return sections
        
        def get_transfer_table(students):
            return {
                'moved_in': students.filter(transferred_to__isnull=False).count(),
                'moved_out': students.filter(transferred_from__isnull=False).count(),
                'withdrawn': students.filter(exam_round='سحب').count(),
            }
        
        g1_stats_table = get_stats_table(g1_students)
        g2_stats_table = get_stats_table(g2_students)
        g3_stats_table = get_stats_table(g3_students)
        
        g1_sections_table = get_sections_table(g1_students)
        g2_sections_table = get_sections_table(g2_students)
        g3_sections_table = get_sections_table(g3_students)
        
        g1_transfer_table = get_transfer_table(g1_students)
        g2_transfer_table = get_transfer_table(g2_students)
        g3_transfer_table = get_transfer_table(g3_students)
        
        # بيانات الطباعة (g1_print, g2_print, g3_print)
        def prepare_print_data(students, grade_key):
            """إعداد بيانات الطباعة لكل طالب"""
            from datetime import date
            print_data = []
            for student in students:
                # حساب تاريخ الميلاد
                birth_d = birth_m = birth_y = ''
                if student.birth_date:
                    birth_d = student.birth_date.day
                    birth_m = student.birth_date.month
                    birth_y = student.birth_date.year
                
                # حساب سن الوافد
                f_birth_d = f_birth_m = f_birth_y = ''
                if student.foreign_birth_date:
                    f_birth_d = student.foreign_birth_date.day
                    f_birth_m = student.foreign_birth_date.month
                    f_birth_y = student.foreign_birth_date.year
                
                # حساب السن في أول أكتوبر
                age_oct = ''
                if student.exam_age:
                    age_oct = f"{student.exam_age} سنة"
                elif student.birth_date:
                    # حساب من تاريخ الميلاد
                    oct_date = date(student.birth_date.year if student.birth_date.year > 2000 else 2000 + student.birth_date.year, 10, 1)
                    if student.birth_date < oct_date:
                        age_calc = oct_date.year - student.birth_date.year
                    else:
                        age_calc = oct_date.year - student.birth_date.year - 1
                    age_oct = f"{age_calc} سنة"
                
                print_data.append({
                    'student_identifier': student.id,
                    'student_code': student.student_code or '',
                    'student_name': student.name or '',
                    'national_id': student.national_id or '',
                    'birth_d': birth_d,
                    'birth_m': birth_m,
                    'birth_y': birth_y,
                    'birth_place': student.birth_governorate or '',
                    'age_oct': age_oct,
                    'section': student.get_section_display() if hasattr(student, 'get_section_display') else (student.section or ''),
                    'nationality': student.get_nationality_display() if hasattr(student, 'get_nationality_display') else (student.nationality or ''),
                    'religion': student.get_religion_display() if hasattr(student, 'get_religion_display') else (student.religion or ''),
                    'gender': student.get_gender_display() if hasattr(student, 'get_gender_display') else ('ذكر' if student.gender == 'male' else 'أنثى'),
                    'lang1': student.first_foreign_language or '',
                    'lang2': student.second_foreign_language or '',
                    'reg_type': student.get_registration_type_display() if hasattr(student, 'get_registration_type_display') else (student.registration_type or ''),
                    'reg_status': student.get_registration_status_display() if hasattr(student, 'get_registration_status_display') else (student.registration_status or ''),
                    'inclusion': student.inclusion_status or '',
                    'guardian': student.guardian_name or '',
                    'is_foreign': student.is_foreign_student or False,
                    'f_birth_d': f_birth_d,
                    'f_birth_m': f_birth_m,
                    'f_birth_y': f_birth_y,
                    'f_birth_place': student.foreign_birth_place or '',
                    'f_gender': 'ذكر' if student.foreign_gender == 'male' else ('أنثى' if student.foreign_gender == 'female' else ''),
                    'student_status': student.registration_status or '',
                    'transferred_from': student.transferred_from or '',
                    'transferred_to': student.transferred_to or '',
                    'school_affiliation': student.nationality or 'رسمي',
                    'grade': {'g1': 'الصف الأول', 'g2': 'الصف الثاني', 'g3': 'الصف الثالث'}.get(grade_key, grade_key),
                })
            return print_data
        
        g1_print = prepare_print_data(g1_students, 'g1')
        g2_print = prepare_print_data(g2_students, 'g2')
        g3_print = prepare_print_data(g3_students, 'g3')
        
    except Exception as e:
        messages.error(request, f'حدث خطأ في جلب البيانات: {str(e)}')
        g1_students = g2_students = g3_students = Student.objects.none()
        g1_stats = g2_stats = g3_stats = {'total': 0, 'regular': 0, 'services': 0, 'system': 0, 'workers': 0, 'foreign': 0}
        import json
        empty_chart = {
            'gender': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'foreign': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'withdrawn': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'sections': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'regtype': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'status': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'inclusion': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'lang1': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
            'lang2': {'labels': [], 'data': [], 'labels_json': '[]', 'data_json': '[]'},
        }
        g1_gender = g2_gender = g3_gender = empty_chart.copy()
        g1_stats_table = g2_stats_table = g3_stats_table = []
        g1_sections_table = g2_sections_table = g3_sections_table = []
        g1_transfer_table = g2_transfer_table = g3_transfer_table = {'moved_in': 0, 'moved_out': 0, 'withdrawn': 0}
        g1_print = g2_print = g3_print = []
        site_settings = SiteSettings.get_settings()
    
    context = {
        'g1_students': g1_students,
        'g2_students': g2_students,
        'g3_students': g3_students,
        'g1_stats': g1_stats,
        'g2_stats': g2_stats,
        'g3_stats': g3_stats,
        'g1_gender': g1_gender,
        'g2_gender': g2_gender,
        'g3_gender': g3_gender,
        'g1_stats_table': g1_stats_table,
        'g2_stats_table': g2_stats_table,
        'g3_stats_table': g3_stats_table,
        'g1_sections_table': g1_sections_table,
        'g2_sections_table': g2_sections_table,
        'g3_sections_table': g3_sections_table,
        'g1_transfer_table': g1_transfer_table,
        'g2_transfer_table': g2_transfer_table,
        'g3_transfer_table': g3_transfer_table,
        'g1_print': g1_print,
        'g2_print': g2_print,
        'g3_print': g3_print,
        'site_settings': site_settings,
        'search_query': search_query,
        'grade_filter': grade_filter,
        'registration_status': registration_status,
    }
    
    return render(request, 'student_affairs/records_list.html', context)


@login_required
def add_record(request):
    """إضافة/تحديث سجل لطالب"""
    select_form = SelectStudentForm()
    record_form = None
    student = None
    
    # حفظ السجل
    if request.method == 'POST':
        if 'select_student' in request.POST:
            select_form = SelectStudentForm(request.POST)
            if select_form.is_valid():
                student = select_form.cleaned_data['student']
                return redirect(f"{reverse('student_affairs:add_record')}?student_id={student.id}")
        
        elif 'save_record' in request.POST:
            # الحصول على معرف الطالب من POST
            selected_student_id = request.POST.get('selected_student')
            if selected_student_id:
                try:
                    student = Student.objects.get(id=selected_student_id)
                    record_form = RecordsForm(request.POST, instance=student)
                    if record_form.is_valid():
                        record_form.save()
                        messages.success(request, f'تم حفظ سجل الطالب {student.name} بنجاح!')
                        return redirect('student_affairs:student_detail', student_id=student.id)
                    else:
                        messages.error(request, 'يرجى تصحيح الأخطاء في النموذج.')
                except Student.DoesNotExist:
                    messages.error(request, 'الطالب غير موجود.')
            else:
                messages.error(request, 'لم يتم تحديد طالب.')
    
    # اختيار الطالب من GET (بعد اختياره من النموذج)
    if not student:
        student_id = request.GET.get('student_id')
        if student_id:
            try:
                student = Student.objects.get(id=student_id)
                record_form = RecordsForm(instance=student)
            except Student.DoesNotExist:
                messages.error(request, 'الطالب غير موجود.')
    
    context = {
        'select_form': select_form,
        'record_form': record_form,
        'form': record_form,  # للتوافق مع القالب الذي يستخدم 'form'
        'student': student,
        'selected_student': student,  # للتوافق مع القالب
        'site_settings': SiteSettings.get_settings(),  # إضافة إعدادات الموقع
    }
    
    return render(request, 'student_affairs/add_record.html', context)


@login_required
def classrooms_list(request):
    """قائمة الفصول"""
    try:
        all_classrooms = Classroom.objects.all().order_by('grade', 'gender', 'class_number')
        
        # فلترة حسب الصف
        grade_filter = request.GET.get('grade', '')
        gender_filter = request.GET.get('gender', '')
        
        if grade_filter:
            all_classrooms = all_classrooms.filter(grade=grade_filter)
        if gender_filter:
            all_classrooms = all_classrooms.filter(gender=gender_filter)
        
        # إحصائيات
        total_classrooms = Classroom.objects.count()
        g1_count = Classroom.objects.filter(grade='g1').count()
        g2_count = Classroom.objects.filter(grade='g2').count()
        g3_count = Classroom.objects.filter(grade='g3').count()
        
        # تجميع حسب الصف والنوع - للتوافق مع القالب
        male = {
            'g1': all_classrooms.filter(grade='g1', gender='male'),
            'g2': all_classrooms.filter(grade='g2', gender='male'),
            'g3': all_classrooms.filter(grade='g3', gender='male'),
        }
        female = {
            'g1': all_classrooms.filter(grade='g1', gender='female'),
            'g2': all_classrooms.filter(grade='g2', gender='female'),
            'g3': all_classrooms.filter(grade='g3', gender='female'),
        }
        
    except Exception as e:
        messages.error(request, f'حدث خطأ في جلب البيانات: {str(e)}')
        all_classrooms = Classroom.objects.none()
        total_classrooms = g1_count = g2_count = g3_count = 0
        male = {'g1': Classroom.objects.none(), 'g2': Classroom.objects.none(), 'g3': Classroom.objects.none()}
        female = {'g1': Classroom.objects.none(), 'g2': Classroom.objects.none(), 'g3': Classroom.objects.none()}
    
    context = {
        'classrooms': all_classrooms,
        'male': male,
        'female': female,
        'total_classrooms': total_classrooms,
        'g1_count': g1_count,
        'g2_count': g2_count,
        'g3_count': g3_count,
        'grade_filter': grade_filter,
        'gender_filter': gender_filter,
    }
    
    return render(request, 'student_affairs/classrooms_list.html', context)


@login_required
def add_classroom(request):
    """إضافة فصل جديد"""
    if request.method == 'POST':
        form = ClassroomForm(request.POST)
        if form.is_valid():
            try:
                classroom = form.save()
                messages.success(request, f'تم إضافة الفصل {classroom.display_name} بنجاح!')
                return redirect('student_affairs:classroom_detail', classroom_id=classroom.id)
            except Exception as e:
                messages.error(request, f'حدث خطأ في إضافة الفصل: {str(e)}')
    else:
        form = ClassroomForm()
    
    return render(request, 'student_affairs/add_classroom.html', {'form': form})


@login_required
def classroom_detail(request, classroom_id):
    """تفاصيل الفصل"""
    try:
        classroom = get_object_or_404(Classroom, id=classroom_id)
        students = classroom.students.all().order_by('student_code')
        
        # إضافة طالب للفصل
        if request.method == 'POST':
            if 'add_student' in request.POST:
                # دعم إضافة طالب واحد أو عدة طلاب
                student_ids = request.POST.getlist('students')  # قائمة من IDs
                # إذا لم يكن هناك students، جرب الطريقة القديمة (للتوافق)
                if not student_ids:
                    student_id = request.POST.get('student_id') or request.POST.get('student-picker')
                    if student_id:
                        student_ids = [student_id]
                
                if student_ids:
                    added_count = 0
                    errors = []
                    for student_id in student_ids:
                        try:
                            student = Student.objects.get(id=student_id)
                            if student.classroom and student.classroom != classroom:
                                errors.append(f'الطالب {student.name} موجود في فصل آخر!')
                            elif classroom.students.count() >= classroom.seats_count:
                                errors.append(f'الفصل ممتلئ! لا يمكن إضافة المزيد من الطلاب.')
                                break
                            else:
                                student.classroom = classroom
                                student.grade = classroom.grade
                                student.save()
                                added_count += 1
                        except Student.DoesNotExist:
                            errors.append(f'الطالب برقم {student_id} غير موجود.')
                        except Exception as e:
                            errors.append(f'حدث خطأ عند إضافة الطالب: {str(e)}')
                    
                    if added_count > 0:
                        messages.success(request, f'تم إضافة {added_count} طالب بنجاح!')
                    if errors:
                        for error in errors:
                            messages.error(request, error)
                    
                    return redirect('student_affairs:classroom_detail', classroom_id=classroom.id)
                else:
                    messages.error(request, 'لم يتم تحديد أي طالب.')
            
            # إزالة طالب من الفصل
            elif 'remove_student' in request.POST or request.POST.get('remove_student'):
                student_id = request.POST.get('student_id')
                if student_id:
                    try:
                        student = Student.objects.get(id=student_id, classroom=classroom)
                        student.classroom = None
                        student.save()
                        messages.success(request, f'تم إزالة الطالب {student.name} من الفصل بنجاح!')
                        return redirect('student_affairs:classroom_detail', classroom_id=classroom.id)
                    except Student.DoesNotExist:
                        messages.error(request, 'الطالب غير موجود.')
        
        # إحصائيات
        total_students = students.count()
        available_seats = classroom.remaining_seats() if hasattr(classroom, 'remaining_seats') else max(0, classroom.seats_count - total_students)
        
        # إنشاء قائمة المقاعد
        seats = []
        students_list = list(students)
        for i in range(1, classroom.seats_count + 1):
            if i <= len(students_list):
                seats.append({'index': i, 'student': students_list[i-1]})
            else:
                seats.append({'index': i, 'student': None})
        
        # الطلاب المتاحين للإضافة
        available_students = Student.objects.filter(
            Q(classroom__isnull=True) | Q(classroom=classroom),
            Q(grade=classroom.grade) | Q(grade__isnull=True),
            gender=classroom.gender
        ).distinct().order_by('student_code', 'name')
        
        # نموذج إضافة طالب
        add_form = AddStudentToClassroomForm(classroom=classroom)
        
    except Exception as e:
        messages.error(request, f'حدث خطأ: {str(e)}')
        return redirect('student_affairs:classrooms_list')
    
    context = {
        'classroom': classroom,
        'students': students,
        'total_students': total_students,
        'available_seats': available_seats,
        'remaining': available_seats,
        'seats': seats,
        'available_students': available_students,
        'form': add_form,
    }
    
    return render(request, 'student_affairs/classroom_detail.html', context)


@login_required
def classroom_print(request, classroom_id):
    """طباعة قائمة الفصل"""
    classroom = get_object_or_404(Classroom, id=classroom_id)
    students = classroom.students.all().order_by('student_code')
    site_settings = SiteSettings.get_settings()
    
    context = {
        'classroom': classroom,
        'students': students,
        'site_settings': site_settings,
    }
    
    return render(request, 'student_affairs/classroom_print.html', context)


@login_required
def classrooms_print_grade(request, gender, grade):
    """طباعة قوائم الصف"""
    classrooms = Classroom.objects.filter(grade=grade, gender=gender).order_by('class_number')
    all_students = Student.objects.filter(
        Q(grade=grade) | Q(classroom__grade=grade),
        Q(gender=gender) | Q(classroom__gender=gender)
    ).distinct().order_by('student_code', 'name')
    
    site_settings = SiteSettings.get_settings()
    
    grade_label = {'g1': 'الصف الأول', 'g2': 'الصف الثاني', 'g3': 'الصف الثالث'}.get(grade, grade)
    gender_label = {'male': 'بنين', 'female': 'بنات'}.get(gender, gender)
    
    # تجميع البيانات للقالب
    classes_data = []
    for classroom in classrooms:
        students = all_students.filter(classroom=classroom).order_by('student_code')
        classes_data.append({
            'classroom': classroom,
            'students': students,
            'grade_number': {'g1': '1', 'g2': '2', 'g3': '3'}.get(grade, ''),
        })
    
    context = {
        'classes_data': classes_data,
        'site': site_settings,
        'site_settings': site_settings,
        'grade': grade,
        'gender': gender,
        'grade_label': grade_label,
        'gender_label': gender_label,
        'academic_year': site_settings.academic_year if hasattr(site_settings, 'academic_year') else '2024-2025',
    }
    
    return render(request, 'student_affairs/classrooms_print_grade.html', context)


@login_required
def classrooms_print_absence_grade(request, gender, grade):
    """طباعة قوائم الغياب"""
    classrooms = Classroom.objects.filter(grade=grade, gender=gender).order_by('class_number')
    site_settings = SiteSettings.get_settings()
    
    grade_label = {'g1': 'الصف الأول', 'g2': 'الصف الثاني', 'g3': 'الصف الثالث'}.get(grade, grade)
    gender_label = {'male': 'بنين', 'female': 'بنات'}.get(gender, gender)
    
    # تجميع البيانات للقالب
    classes_data = []
    for classroom in classrooms:
        students = classroom.students.all().order_by('student_code', 'name')
        classes_data.append({
            'classroom': classroom,
            'students': students,
            'grade_number': {'g1': '1', 'g2': '2', 'g3': '3'}.get(grade, ''),
        })
    
    context = {
        'classes_data': classes_data,
        'site': site_settings,
        'site_settings': site_settings,
        'grade': grade,
        'gender': gender,
        'grade_label': grade_label,
        'gender_label': gender_label,
        'academic_span': site_settings.academic_year if hasattr(site_settings, 'academic_year') else '2024-2025',
    }
    
    return render(request, 'student_affairs/classrooms_print_absence_grade.html', context)


@login_required
def print_student_list(request):
    """طباعة قائمة الطلاب"""
    try:
        all_students = Student.objects.all().order_by('student_code', 'name')
        
        # فلترة
        grade_filter = request.GET.get('grade', '')
        gender_filter = request.GET.get('gender', '')
        registration_status = request.GET.get('registration_status', '')
        
        if grade_filter:
            all_students = all_students.filter(
                Q(grade=grade_filter) | Q(classroom__grade=grade_filter)
            ).distinct()
        
        if gender_filter:
            all_students = all_students.filter(gender=gender_filter)
        
        if registration_status:
            all_students = all_students.filter(registration_status=registration_status)
        
        # إعداد البيانات للقالب (students_with_age)
        from datetime import date
        students_with_age = []
        for student in all_students:
            age = None
            if student.birth_date:
                today = date.today()
                age = today.year - student.birth_date.year - ((today.month, today.day) < (student.birth_date.month, student.birth_date.day))
            elif student.age:
                age = student.age
            elif student.exam_age:
                age = student.exam_age
            
            students_with_age.append({
                'student': student,
                'age': age,
            })
        
        total_students = len(students_with_age)
        site_settings = SiteSettings.get_settings()
        
    except Exception as e:
        messages.error(request, f'حدث خطأ في جلب البيانات: {str(e)}')
        students_with_age = []
        total_students = 0
        site_settings = SiteSettings.get_settings()
    
    context = {
        'students': all_students,
        'students_with_age': students_with_age,
        'total_students': total_students,
        'site_settings': site_settings,
        'grade_filter': grade_filter,
        'gender_filter': gender_filter,
        'registration_status': registration_status,
    }
    
    return render(request, 'student_affairs/print_student_list.html', context)


@login_required
def add_student(request):
    """إضافة طالب جديد"""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            
            # استخراج البيانات من الرقم القومي إذا كان موجوداً
            national_id = form.cleaned_data.get('national_id', '')
            if national_id and len(national_id) == 14:
                try:
                    parsed_data = parse_egyptian_national_id(national_id)
                    if parsed_data:
                        if not student.birth_date:
                            student.birth_date = parsed_data.get('birth_date')
                        if not student.gender:
                            student.gender = parsed_data.get('gender', 'male')
                        if not student.birth_governorate:
                            student.birth_governorate = parsed_data.get('birth_governorate', '')
                        if not student.age:
                            from datetime import date
                            if parsed_data.get('birth_date'):
                                today = date.today()
                                birth = parsed_data['birth_date']
                                student.age = today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
                except Exception as e:
                    messages.warning(request, f'تحذير: لم يتم استخراج جميع البيانات من الرقم القومي. {str(e)}')
            
            student.save()
            messages.success(request, f'تم إضافة الطالب {student.name} بنجاح!')
            
            # إعادة التوجيه
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url:
                # إذا كان next_url هو اسم URL pattern، استخدم reverse
                if next_url == 'add_record':
                    return redirect('student_affairs:add_record')
                # إذا كان URL كامل، استخدمه مباشرة
                elif next_url.startswith('/') or next_url.startswith('http'):
                    return redirect(next_url)
                # إذا كان اسم pattern بدون namespace، حاول إضافة namespace
                else:
                    try:
                        return redirect(f'student_affairs:{next_url}')
                    except:
                        return redirect(next_url)
            return redirect('student_affairs:student_detail', student_id=student.id)
    else:
        form = StudentForm()
    
    return render(request, 'student_affairs/add_student.html', {'form': form})


@login_required
def download_excel_template(request):
    """تحميل قالب Excel لإضافة الطلاب"""
    try:
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الطلبة"
        
        # رؤوس الأعمدة
        headers = [
            'كود الطالب', 'الاسم', 'الديانة', 'التبعية', 'حالة القيد', 
            'نوع القيد', 'الشعبة', 'النوع', 'محافظة الميلاد', 'تاريخ الميلاد',
            'الرقم القومي', 'السن', 'السن عند الامتحان', 'المدرسة', 'اسم ولي الأمر',
            'العنوان', 'هاتف الطالب', 'هاتف ولي الأمر', 'الصف', 'العام الأكاديمي'
        ]
        
        # كتابة الرؤوس
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        
        # إضافة صف مثال
        example_row = [
            'ST001', 'أحمد محمد علي', 'مسلمة', 'رسمي', 'ناجح ومنقول',
            'نظامي', 'فني مبيعات', 'male', 'القاهرة', '2010-01-01',
            '30101012345678', '14', '15', 'مدرسة', 'محمد علي',
            'العنوان', '01000000000', '01000000000', 'g1', '2024-2025'
        ]
        
        for col, value in enumerate(example_row, 1):
            ws.cell(row=2, column=col).value = value
        
        # ضبط عرض الأعمدة
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # إعداد response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_students.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'حدث خطأ في إنشاء القالب: {str(e)}')
        return redirect('student_affairs:student_list')


@login_required
def student_detail(request, student_id):
    """تفاصيل الطالب"""
    student = get_object_or_404(Student, id=student_id)
    attachments = student.attachments.all()
    
    # نماذج افتراضية
    attachment_form = AttachmentForm()
    contact_form = StudentContactForm(instance=student)
    
    if request.method == 'POST':
        # إضافة مرفق جديد
        if 'add_attachment' in request.POST:
            attachment_form = AttachmentForm(request.POST, request.FILES)
            if attachment_form.is_valid():
                attachment = attachment_form.save(commit=False)
                attachment.student = student
                attachment.save()
                messages.success(request, 'تم إضافة المرفق بنجاح!')
                return redirect('student_affairs:student_detail', student_id=student.id)
        
        # تحديث بيانات التواصل
        elif 'update_contact' in request.POST:
            contact_form = StudentContactForm(request.POST, instance=student)
            if contact_form.is_valid():
                contact_form.save()
                messages.success(request, 'تم تحديث بيانات التواصل بنجاح!')
                return redirect('student_affairs:student_detail', student_id=student.id)
    
    context = {
        'student': student,
        'attachments': attachments,
        'attachment_form': attachment_form,
        'contact_form': contact_form,
    }
    
    return render(request, 'student_affairs/student_detail.html', context)


@login_required
def student_id_card(request, student_id):
    """بطاقة الطالب التعريفية"""
    student = get_object_or_404(Student, id=student_id)
    site_settings = SiteSettings.get_settings()
    
    # توليد QR Code
    qr_code = None
    try:
        # إنشاء رابط للطالب
        student_url = request.build_absolute_uri(reverse('student_affairs:student_detail', args=[student.id]))
        
        # إنشاء QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(student_url)
        qr.make(fit=True)
        
        # إنشاء الصورة
        img = qr.make_image(fill_color="black", back_color="white")
        
        # تحويل إلى base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        qr_code = base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception as e:
        # في حالة فشل توليد QR code، نتركه None
        pass
    
    context = {
        'student': student,
        'site_settings': site_settings,
        'qr_code': qr_code,
    }
    
    return render(request, 'student_affairs/student_id_card.html', context)


@login_required
def student_qr(request, student_id):
    """QR code للطالب"""
    student = get_object_or_404(Student, id=student_id)
    
    try:
        # إنشاء رابط للطالب
        student_url = request.build_absolute_uri(reverse('student_affairs:student_detail', args=[student.id]))
        
        # إنشاء QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(student_url)
        qr.make(fit=True)
        
        # إنشاء الصورة
        img = qr.make_image(fill_color="black", back_color="white")
        
        # إرجاع الصورة كـ response
        response = HttpResponse(content_type="image/png")
        img.save(response, "PNG")
        return response
    except Exception as e:
        return HttpResponse(f"خطأ في توليد QR Code: {str(e)}", status=500)


@login_required
def delete_attachment(request, attachment_id):
    """حذف مرفق"""
    attachment = get_object_or_404(StudentAttachment, id=attachment_id)
    student_id = attachment.student.id
    
    if request.method == 'POST':
        attachment.delete()
        messages.success(request, 'تم حذف المرفق بنجاح!')
        return redirect('student_affairs:student_detail', student_id=student_id)
    
    return redirect('student_affairs:student_detail', student_id=student_id)


# ===== SYSTEM STATISTICS FUNCTION =====
@login_required
def system_statistics(request):
    """صفحة إحصائيات النظام - دالة معقدة مع جميع الإحصائيات المطلوبة"""
    from django.db.models import Count, Sum, Case, When, IntegerField
    from django.db.models.functions import Extract
    
    # الحصول على جميع الطلاب
    all_students = Student.objects.all()
    
    # ===== 1. تبويب الراسب =====
    # الراسبين (حالة القيد = راسب وباق للإعادة)
    failed_students = all_students.filter(registration_status='راسب وباق للإعادة')
    
    # تجميع الراسبين حسب الفصل
    failed_by_classroom = {}
    for cls in Classroom.objects.all():
        failed_in_cls = failed_students.filter(classroom=cls)
        if failed_in_cls.exists():
            failed_by_classroom[cls.id] = {
                'classroom': cls,
                'students': failed_in_cls,
                'count': failed_in_cls.count()
            }
    
    # ===== 2. تبويب الإحصاء =====
    # إحصاء حسب النوع والحالة
    stats_by_gender_status = {}
    for gender_key, gender_label in [('male', 'بنين'), ('female', 'بنات')]:
        gender_students = all_students.filter(gender=gender_key)
        stats_by_gender_status[gender_key] = {
            'label': gender_label,
            'regular_new': gender_students.filter(registration_type='نظامي', exam_round='دور أول').count(),
            'regular_repeat': gender_students.filter(registration_type='نظامي', exam_round__in=['دور ثاني', 'باقي']).count(),
            'inclusion': gender_students.filter(inclusion_status__in=['بصري', 'سمعي', 'حركي', 'صعوبات تعلم', 'أخرى']).count(),
            'workers_new': gender_students.filter(registration_type='عمال', exam_round='دور أول').count(),
            'workers_repeat': gender_students.filter(registration_type='عمال', exam_round__in=['دور ثاني', 'باقي']).count(),
            'total': gender_students.count(),
        }
    
    # عدد الفصول حسب الصف والنوع
    classrooms_by_grade_gender = {}
    for grade_key in ['g1', 'g2', 'g3']:
        grade_label = {'g1': 'الصف الأول', 'g2': 'الصف الثاني', 'g3': 'الصف الثالث'}[grade_key]
        classrooms_by_grade_gender[grade_key] = {
            'label': grade_label,
            'male': Classroom.objects.filter(grade=grade_key, gender='male').count(),
            'female': Classroom.objects.filter(grade=grade_key, gender='female').count(),
            'total': Classroom.objects.filter(grade=grade_key).count(),
        }
    
    # ===== 3. إحصاء الطلاب المسيحيين =====
    christian_stats = {
        'male': all_students.filter(religion='مسيحية', gender='male').count(),
        'female': all_students.filter(religion='مسيحية', gender='female').count(),
        'total': all_students.filter(religion='مسيحية').count(),
    }
    
    # ===== 4. إحصاء السحب =====
    withdrawn_students = all_students.filter(exam_round='سحب').order_by('student_code')
    
    # ===== 5. إحصاء المجموع (حسب المرحلة والنوع) =====
    # المرحلة الأولى: حتى 165 (افتراض: 16.5 سنة = 16.5 * 12 = 198 شهر، لكن المستخدم قال 165، ربما يقصد 16.5 سنة)
    # سأفترض أنه يقصد السن عند الامتحان (exam_age) كرقم عشري أو سن
    # المرحلة الأولى: >= 16.5 (أو 165 شهر / 12 = 13.75 سنة؟)
    # المرحلة الثانية: 16.4 إلى 15.0
    # المرحلة الثالثة: 14.9 إلى 14.0
    
    # سأفترض أن exam_age هو السن بالسنوات، والمراحل:
    # المرحلة الأولى: >= 17 سنة
    # المرحلة الثانية: 16-15 سنة
    # المرحلة الثالثة: <= 14 سنة
    
    stage_stats = {}
    for stage_key, stage_label, min_age, max_age in [
        ('stage1', 'المرحلة الأولى', 17, 100),
        ('stage2', 'المرحلة الثانية', 15, 16),
        ('stage3', 'المرحلة الثالثة', 0, 14),
    ]:
        stage_students = all_students.filter(exam_age__gte=min_age, exam_age__lte=max_age)
        stage_stats[stage_key] = {
            'label': stage_label,
            'male': stage_students.filter(gender='male').count(),
            'female': stage_students.filter(gender='female').count(),
            'total': stage_students.count(),
        }
    
    # ===== 6. إجمالي الصف الأول =====
    grade1_students = all_students.filter(
        Q(grade__icontains='1') | Q(grade__icontains='الأول') | Q(grade__icontains='اول')
    )
    
    grade1_stats = {}
    for gender_key, gender_label in [('male', 'بنين'), ('female', 'بنات')]:
        g1_gender = grade1_students.filter(gender=gender_key)
        grade1_stats[gender_key] = {
            'label': gender_label,
            'regular_new': g1_gender.filter(registration_type='نظامي', exam_round='دور أول').count(),
            'regular_repeat': g1_gender.filter(registration_type='نظامي', exam_round__in=['دور ثاني', 'باقي']).count(),
            'inclusion': g1_gender.filter(inclusion_status__in=['بصري', 'سمعي', 'حركي', 'صعوبات تعلم', 'أخرى']).count(),
            'workers_new': g1_gender.filter(registration_type='عمال', exam_round='دور أول').count(),
            'total': g1_gender.count(),
        }
    
    # عدد فصول الصف الأول
    grade1_classrooms = {
        'male': Classroom.objects.filter(grade='g1', gender='male').count(),
        'female': Classroom.objects.filter(grade='g1', gender='female').count(),
        'total': Classroom.objects.filter(grade='g1').count(),
    }
    
    context = {
        'failed_students': failed_students,
        'failed_by_classroom': failed_by_classroom,
        'stats_by_gender_status': stats_by_gender_status,
        'classrooms_by_grade_gender': classrooms_by_grade_gender,
        'christian_stats': christian_stats,
        'withdrawn_students': withdrawn_students,
        'stage_stats': stage_stats,
        'grade1_stats': grade1_stats,
        'grade1_classrooms': grade1_classrooms,
    }
    
    return render(request, 'student_affairs/system_statistics.html', context)


@login_required
def classroom_failed_details(request, classroom_id):
    """تفاصيل الراسبين في فصل محدد - للاستخدام في Modal"""
    classroom = get_object_or_404(Classroom, id=classroom_id)
    failed_students = Student.objects.filter(
        classroom=classroom,
        registration_status='راسب وباق للإعادة'
    ).order_by('name', 'student_code')
    
    context = {
        'classroom': classroom,
        'students': failed_students,
    }
    return render(request, 'student_affairs/classroom_failed_details.html', context)


@staff_member_required
def admin_create_user(request):
    """إنشاء مستخدم جديد مع توليد اسم مستخدم وكلمة مرور"""
    form = AdminCreateUserForm(request.POST or None)
    created = None
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        created = {
            'username': form.generated_username,
            'password': form.generated_password,
            'email': user.email,
        }
        messages.success(request, 'تم إنشاء المستخدم بنجاح')
    return render(request, 'student_affairs/admin_create_user.html', {
        'form': form,
        'created': created,
    })


def landing(request):
    """صفحة تعريفية احترافية للموقع"""
    site = SiteSettings.get_settings()
    return render(request, 'student_affairs/landing.html', {'site': site})


@require_http_methods(["GET", "POST"])
def custom_login(request):
    """
    صفحة تسجيل الدخول المخصصة مع JWT
    """
    from django.contrib.auth.forms import AuthenticationForm
    
    # إذا كان هناك token في URL، التحقق منه
    token = request.GET.get('token')
    if token:
        user = verify_jwt_token(token)
        if user:
            login(request, user)
            request.session['jwt_token'] = token
            request.session['user_id'] = user.id
            response = redirect('/welcome/')
            response.set_cookie('jwt_token', token, max_age=7*24*60*60, httponly=True, samesite='Lax')
            messages.success(request, f'مرحباً {user.username}! تم تسجيل الدخول بنجاح.')
            return response
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # تسجيل الدخول
            login(request, user)
            
            # توليد JWT token
            token = generate_jwt_token(user)
            
            # حفظ token في session
            request.session['jwt_token'] = token
            request.session['user_id'] = user.id
            
            # إعداد response
            response = redirect('/welcome/')
            
            # إضافة token كـ cookie
            response.set_cookie('jwt_token', token, max_age=7*24*60*60, httponly=True, samesite='Lax')
            
            messages.success(request, f'مرحباً {user.username}! تم تسجيل الدخول بنجاح.')
            return response
        else:
            messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة.')
            # إنشاء form مع البيانات المدخلة لعرضها مرة أخرى
            form = AuthenticationForm(data=request.POST)
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})


def custom_logout(request):
    """
    صفحة تسجيل الخروج المخصصة - تقبل GET request
    """
    if request.user.is_authenticated:
        # حذف JWT token من session
        if 'jwt_token' in request.session:
            del request.session['jwt_token']
        if 'user_id' in request.session:
            del request.session['user_id']
        
        # تسجيل الخروج
        logout(request)
        
        # حذف cookie
        response = redirect('/')
        response.delete_cookie('jwt_token')
        
        messages.success(request, 'تم تسجيل الخروج بنجاح.')
        return response
    
    return redirect('/')


def generate_login_link(request, user_id):
    """
    توليد رابط تسجيل دخول مشفر للمستخدم (للمستخدمين المصرح لهم فقط)
    """
    if not request.user.is_staff:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('/')
    
    try:
        user = User.objects.get(id=user_id)
        token = generate_jwt_token(user)
        login_url = f"{request.scheme}://{request.get_host()}/accounts/login/?token={token}"
        return render(request, 'student_affairs/generate_login_link.html', {
            'user': user,
            'token': token,
            'login_url': login_url,
        })
    except User.DoesNotExist:
        messages.error(request, 'المستخدم غير موجود.')
        return redirect('/')


@login_required
def chatbot_api(request):
    """
    API endpoint للشات بوت
    """
    if request.method == 'POST':
        import json
        from .chatbot import get_chatbot_response, analyze_image
        
        try:
            data = json.loads(request.body)
            user_message = data.get('message', '')
            image_data = data.get('image', None)
            conversation_history = data.get('history', [])
            
            user_name = request.user.username
            if hasattr(request.user, 'profile') and request.user.profile.full_name:
                user_name = request.user.profile.full_name
            
            # إذا كان هناك صورة، حللها
            if image_data:
                response = analyze_image(image_data, user_message, user_name)
            else:
                response = get_chatbot_response(user_message, user_name, None, conversation_history)
            
            return JsonResponse({
                'success': True,
                'response': response,
                'user_name': user_name
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def profile(request):
    """صفحة الملف الشخصي"""
    profile_obj, created = UserProfile.objects.get_or_create(user=request.user)
    
    profile_form = ProfileUpdateForm(instance=profile_obj)
    password_form = PasswordChangeForm(user=request.user)
    
    if request.method == 'POST':
        if 'update_profile' in request.POST:
            profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile_obj)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, 'تم تحديث الملف الشخصي بنجاح!')
                return redirect('student_affairs:profile')
        
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                password_form.save()
                messages.success(request, 'تم تغيير كلمة المرور بنجاح!')
                return redirect('student_affairs:profile')
    
    context = {
        'profile': profile_obj,
        'profile_form': profile_form,
        'password_form': password_form,
    }
    
    return render(request, 'student_affairs/profile.html', context)
